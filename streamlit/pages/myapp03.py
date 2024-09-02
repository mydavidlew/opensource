import streamlit as st
import pandas as pd
import openpyxl as xl
import os

st.set_page_config(page_title="Application #03", page_icon="🌻", layout="wide")
st.sidebar.title("🌻 Application #03")
st.sidebar.markdown(
    """This demo illustrates a combination of geospatial data visualisation, plotting and animation with 
    [**Streamlit**](https://docs.streamlit.io/develop/api-reference). We're generating a bunch of random numbers 
    in a loop for around 5 seconds. Enjoy!"""
)

# Function to get list of worksheet in xlsx
@st.cache_data
def get_xlsx_sheetname(xlsx_file_path):
    with st.spinner('Scanning excel, Wait for it...'):
        excel_file = pd.ExcelFile(xlsx_file_path, engine="calamine")
    st.success("Done! --> " + xlsx_file_path.name)
    # for sheet in excel_file.sheet_names:
    #     ws = pd.read_excel(xlsx_file_path, sheet_name=sheet)
    #     row_count = len(ws.index)
    #     column_count = len(ws.columns)
    #     st.write("sheet:", sheet, "-- row:", row_count, "-- col:", column_count)
    return excel_file.sheet_names

# Function to convert xlsx to csv
def convert_xlsx_to_csv(xlsx_file_path, csv_file_path, sheet_name=0):
    # Read the Excel file
    df = pd.read_excel(xlsx_file_path, sheet_name=sheet_name, header=0)
    # Write the DataFrame to a CSV file
    df.to_csv(csv_file_path, index=False)

# Function to read xlsx and filter distinct values in a specified range
def filter_distinct_values(xlsx_file_path, sheet_name=0, start_row=0, end_row=None, start_col=0, end_col=None):
    # Read the Excel file
    df = pd.read_excel(xlsx_file_path, sheet_name=sheet_name, header=0)
    # Select the specific range
    df_range = df.iloc[start_row:end_row, start_col:end_col]
    # Drop duplicates to filter distinct values
    distinct_values = df_range.drop_duplicates()
    return distinct_values

def domain(xlsx_file_path, csv_file_path):
    tab01, tab02, tab03, tab04, tab05 = st.tabs(["👻 Select worksheet", "👻 Read worksheet", "👻 Convert worksheet", "👻 Filter worksheet", "👻 Other worksheet"])
    sheet = 0
    with tab01:
        #if "df" not in st.session_state:
        #    st.session_state.df = get_xlsx_sheetname(xlsx_file_path)
        #items = st.dataframe(st.session_state.df, use_container_width=True, hide_index=False, on_select="rerun", selection_mode="single-row")
        df = get_xlsx_sheetname(xlsx_file_path)
        items = st.dataframe(df, use_container_width=True, hide_index=False, on_select="rerun", selection_mode="single-row")
        try:
            sheet = items.selection.rows[0]
        except IndexError:
            st.markdown(":red[***No***] :blue[*worksheet is selected!*]")
        st.write("Selected worksheet is: ", sheet)
        csv_file_path = csv_file_path + "_" + str(sheet) + ".csv"

    with tab02:
        nrcds = st.slider('#of Records', min_value=0, max_value=10000, value=100, step=10)
        with st.spinner('Reading excel, Wait for it...'):
            df = pd.read_excel(xlsx_file_path, sheet_name=sheet, header=0, nrows=nrcds)
        st.success("Done! --> " + xlsx_file_path.name)
        st.dataframe(df, use_container_width=True)
        st.write("Total records: ", len(df.index))

    with tab03:
        with st.spinner("Generating file 👉 " + csv_file_path):
            # Convert the file
            convert_xlsx_to_csv(xlsx_file_path, csv_file_path, sheet_name=sheet)
        st.success("Done! --> " + xlsx_file_path.name + " ➡️➡️➡️ " + csv_file_path)

    with tab04:
        # Specify the file path and the range
        start_row = 1  # Adjust these values to your desired range
        end_row = None  # None means until the last row
        start_col = 0  # Column index (0-based)
        end_col = 1  # None  # None means until the last column
        #
        col1, col2, col3, col4 = st.columns([1, 1, 1, 1], gap="large", vertical_alignment="center")
        with col1:
            start_row = st.number_input('start_row', min_value=0, max_value=100, value=1, step=1)
        with col2:
            start_col = st.number_input('start_col', min_value=0, max_value=100, value=0, step=1)
        with col3:
            end_col = st.number_input('end_col', min_value=0, max_value=100, value=1, step=1)
        with col4:
            st.markdown("""
                <style>
                    button {
                        padding-top: 5px !important;
                        padding-bottom: 5px !important;
                        text-align: center !important;
                    }
                </style>
            """, unsafe_allow_html=True)
            bt_start = st.button(r"$\textsf{\Large Start}$", use_container_width=True)
        #
        if bt_start:
            with st.spinner('Filtering, Wait for it...'):
                # Get the distinct values in the specified range
                distinct_values = filter_distinct_values(xlsx_file_path, sheet, start_row, end_row, start_col, end_col)
            st.success("Done! --> " + xlsx_file_path.name)
            # Print or use the distinct values
            st.dataframe(distinct_values, use_container_width=True)

    with tab05:
        with st.spinner('Scanning, Wait for it...'):
            wb = xl.load_workbook(xlsx_file_path, read_only=True)
        st.success("Done! --> " + xlsx_file_path.name)
        sheets = wb.sheetnames
        st.write("worksheets: ", sheets)
        for sheet in sheets:
            ws=wb[sheet]
            row_count = ws.max_row
            column_count = ws.max_column
            st.write(f"sheet: [**{sheet}**] -- row: [**{row_count}**] -- col: [**{column_count}**]")

def main():
    uploaded_excel = st.file_uploader(":blue[**Choose a excel file**]", type=['xls','xlsx'], accept_multiple_files=False)
    if uploaded_excel is not None:
        #st.write(uploaded_excel)
        basefile, extension = os.path.splitext(uploaded_excel.name)
        downloaded_csv = "datasets/" + basefile
        domain(uploaded_excel, downloaded_csv)
    else:
        st.markdown(":red[**Pls upload a excel file...**]")

if __name__ == '__main__':
    st.title("Process xls content")
    main()